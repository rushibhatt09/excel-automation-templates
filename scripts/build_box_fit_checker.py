from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

FONT = "Arial"
wb = Workbook()

BLUE = Font(name=FONT, color="0000FF")
BLACK = Font(name=FONT, color="000000")
GREEN = Font(name=FONT, color="008000")
BOLD = Font(name=FONT, bold=True)
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E5F", end_color="1F4E5F", fill_type="solid")
YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ---------- Read Me ----------
ws = wb.active
ws.title = "Read Me"
ws.sheet_view.showGridLines = False
ws["B2"] = "Volumetric Weight & Box-Fit Checker"
ws["B2"].font = Font(name=FONT, bold=True, size=16, color="1F4E5F")
ws["B4"] = "What this does"
ws["B4"].font = BOLD
ws["B5"] = ("Checks every product against your standard box sizes and tells you the cheapest box it actually fits in - "
            "and flags products that don't fit any box you stock, so you're not overpaying for oversized packaging "
            "or discovering a bad fit at the packing table.")
ws["B7"] = "How to use it"
ws["B7"].font = BOLD
steps = [
    "1. Go to the 'Box Library' tab and replace the sample boxes with your own carton sizes and costs (cells shaded yellow).",
    "2. Go to the 'Product Data' tab and paste your own product dimensions, replacing the sample rows.",
    "3. IMPORTANT: for every product AND every box, enter Length as the longest side, Width as the middle side, and "
    "Height as the shortest side. This keeps the fit-check accurate regardless of which way the box is oriented.",
    "4. Open the 'Fit Checker' tab - it shows the cheapest box each product fits in, or 'NO BOX FITS' if none do.",
]
for i, s in enumerate(steps, start=8):
    ws[f"B{i}"] = s
ws["B13"] = "Notes"
ws["B13"].font = BOLD
ws["B14"] = "- All costs are in INR (Rs), all dimensions in cm, all weights in kg."
ws["B15"] = "- The Box Library should be ordered from smallest/cheapest to largest/costliest - the checker recommends the first (cheapest) box a product fits in."
ws["B16"] = "- This is a generic template with illustrative sample data - it is not tied to any specific brand or product line."
for r in range(5, 17):
    ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws[f"B{r}"].font = Font(name=FONT, size=11)
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 100

# ---------- Box Library ----------
ws = wb.create_sheet("Box Library")
ws.sheet_view.showGridLines = False
ws["A1"] = "Box Library - SAMPLE VALUES, replace with your own carton sizes"
ws["A1"].font = Font(name=FONT, bold=True, size=12, color="1F4E5F")
headers = ["Box Code", "Box Name", "Internal Length (cm)", "Internal Width (cm)", "Internal Height (cm)",
           "Max Weight Capacity (kg)", "Cost per Box (Rs)", "Volume (cm3)"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=3, column=i, value=h)
style_header(ws, 3, len(headers))

box_rows = [
    ["BX1", "XS", 15, 10, 8, 2, 6],
    ["BX2", "S", 20, 15, 10, 5, 8],
    ["BX3", "M", 25, 20, 15, 8, 11],
    ["BX4", "L", 35, 28, 18, 12, 15],
    ["BX5", "XL", 45, 35, 25, 20, 22],
]
for r, row in enumerate(box_rows, start=4):
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = BORDER
        if c in (3, 4, 5, 6, 7):
            cell.font = BLUE
            cell.fill = YELLOW
            if c == 7:
                cell.number_format = "\"Rs\" #,##0"
        else:
            cell.font = BLACK
    vol_cell = ws.cell(row=r, column=8, value=f"=C{r}*D{r}*E{r}")
    vol_cell.font = BLACK
    vol_cell.border = BORDER
    vol_cell.number_format = "#,##0"

ws["A11"] = "Keep boxes ordered smallest/cheapest to largest/costliest - the Fit Checker recommends the first one that fits."
ws["A11"].font = Font(name=FONT, italic=True, size=9, color="666666")
autosize(ws, [10, 10, 18, 16, 16, 20, 16, 14])

# ---------- Product Data ----------
ws = wb.create_sheet("Product Data")
ws.sheet_view.showGridLines = False
pd_headers = ["SKU", "Product Name", "Length (cm)", "Width (cm)", "Height (cm)", "Weight (kg)"]
for i, h in enumerate(pd_headers, start=1):
    ws.cell(row=1, column=i, value=h)
style_header(ws, 1, len(pd_headers))
ws.freeze_panes = "A2"

products = [
    ["SKU-1001", "Face Serum 30ml", 12, 8, 6, 0.15],
    ["SKU-1002", "Sunscreen Combo Pack", 18, 13, 9, 0.35],
    ["SKU-1003", "Face Wash 100ml x2", 16, 11, 7, 0.40],
    ["SKU-1004", "Moisturizer Jar Large", 22, 18, 12, 0.60],
    ["SKU-1005", "Gift Hamper Small", 24, 19, 14, 1.20],
    ["SKU-1006", "Gift Hamper Large", 33, 26, 16, 3.50],
    ["SKU-1007", "Hair Dryer", 30, 14, 10, 0.90],
    ["SKU-1008", "Combo Kit 6-pc", 34, 27, 17, 4.80],
    ["SKU-1009", "Bulk Wholesale Carton", 44, 34, 24, 9.00],
    ["SKU-1010", "Oversized Appliance", 60, 40, 30, 15.00],
]
for r, row in enumerate(products, start=2):
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = BLUE
        cell.border = BORDER
        if c in (3, 4, 5, 6):
            cell.number_format = "0.00"

MAX_ROWS = 200  # Fit Checker/Summary formulas are pre-built this far down so extra pasted rows aren't silently dropped
NOTE_ROW = MAX_ROWS + 10  # kept well beyond MAX_ROWS so it never collides with the blank-row guard on column A

ws[f"A{NOTE_ROW}"] = "Replace with your own product list. Enter Length = longest side, Width = middle side, Height = shortest side."
ws[f"A{NOTE_ROW}"].font = Font(name=FONT, italic=True, size=9, color="666666")
autosize(ws, [14, 26, 14, 12, 14, 12])

LAST_ROW = 1 + MAX_ROWS

# ---------- Fit Checker ----------
ws = wb.create_sheet("Fit Checker")
ws.sheet_view.showGridLines = False
fc_headers = ["SKU", "Product Name", "Length (cm)", "Width (cm)", "Height (cm)", "Weight (kg)",
              "Product Volume (cm3)", "Fits BX1 (XS)", "Fits BX2 (S)", "Fits BX3 (M)", "Fits BX4 (L)",
              "Fits BX5 (XL)", "Recommended Box", "Recommended Box Cost (Rs)", "Space Utilization (%)", "Status"]
for i, h in enumerate(fc_headers, start=1):
    ws.cell(row=1, column=i, value=h)
style_header(ws, 1, len(fc_headers))
ws.freeze_panes = "A2"

box_rows_ref = [4, 5, 6, 7, 8]  # rows in Box Library for BX1..BX5

for i in range(MAX_ROWS):
    r = i + 2
    pd = i + 2
    guard = f"'Product Data'!A{pd}=\"\""

    def blank_if_empty(inner):
        return f'=IF({guard},"",{inner})'

    ws.cell(row=r, column=1, value=blank_if_empty(f"'Product Data'!A{pd}")).font = GREEN
    ws.cell(row=r, column=2, value=blank_if_empty(f"'Product Data'!B{pd}")).font = GREEN
    ws.cell(row=r, column=3, value=blank_if_empty(f"'Product Data'!C{pd}")).font = GREEN
    ws.cell(row=r, column=4, value=blank_if_empty(f"'Product Data'!D{pd}")).font = GREEN
    ws.cell(row=r, column=5, value=blank_if_empty(f"'Product Data'!E{pd}")).font = GREEN
    ws.cell(row=r, column=6, value=blank_if_empty(f"'Product Data'!F{pd}")).font = GREEN
    ws.cell(row=r, column=7, value=blank_if_empty(f"C{r}*D{r}*E{r}")).font = BLACK

    fit_cols = [8, 9, 10, 11, 12]
    for fc, brow in zip(fit_cols, box_rows_ref):
        inner = (f"AND($C{r}<='Box Library'!$C${brow},$D{r}<='Box Library'!$D${brow},"
                 f"$E{r}<='Box Library'!$E${brow},$F{r}<='Box Library'!$F${brow})")
        ws.cell(row=r, column=fc, value=blank_if_empty(inner)).font = GREEN

    H, I, J, K, L = [f"{get_column_letter(c)}{r}" for c in fit_cols]
    rec_box = (f"IF({H},'Box Library'!$A${box_rows_ref[0]},IF({I},'Box Library'!$A${box_rows_ref[1]},"
               f"IF({J},'Box Library'!$A${box_rows_ref[2]},IF({K},'Box Library'!$A${box_rows_ref[3]},"
               f"IF({L},'Box Library'!$A${box_rows_ref[4]},\"NO BOX FITS\")))))")
    ws.cell(row=r, column=13, value=blank_if_empty(rec_box)).font = GREEN

    rec_cost = (f"IF({H},'Box Library'!$G${box_rows_ref[0]},IF({I},'Box Library'!$G${box_rows_ref[1]},"
                f"IF({J},'Box Library'!$G${box_rows_ref[2]},IF({K},'Box Library'!$G${box_rows_ref[3]},"
                f"IF({L},'Box Library'!$G${box_rows_ref[4]},0)))))")
    ws.cell(row=r, column=14, value=blank_if_empty(rec_cost)).font = GREEN

    rec_vol = (f"IF({H},'Box Library'!$H${box_rows_ref[0]},IF({I},'Box Library'!$H${box_rows_ref[1]},"
               f"IF({J},'Box Library'!$H${box_rows_ref[2]},IF({K},'Box Library'!$H${box_rows_ref[3]},"
               f"'Box Library'!$H${box_rows_ref[4]}))))")
    util = f'IF(M{r}="NO BOX FITS",0,G{r}/({rec_vol}))'
    ws.cell(row=r, column=15, value=blank_if_empty(util)).font = BLACK

    status = f'IF(M{r}="NO BOX FITS","REVIEW - NO STANDARD BOX FITS","OK")'
    ws.cell(row=r, column=16, value=blank_if_empty(status)).font = BLACK

    for c in range(1, len(fc_headers) + 1):
        ws.cell(row=r, column=c).border = BORDER
    for c in (3, 4, 5, 6):
        ws.cell(row=r, column=c).number_format = "0.00"
    ws.cell(row=r, column=7).number_format = "#,##0"
    ws.cell(row=r, column=14).number_format = "\"Rs\" #,##0"
    ws.cell(row=r, column=15).number_format = "0.0%"

status_col = "P"
last = LAST_ROW
ws.conditional_formatting.add(f"{status_col}2:{status_col}{last}",
    CellIsRule(operator="equal", formula=['"OK"'], fill=PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")))
ws.conditional_formatting.add(f"{status_col}2:{status_col}{last}",
    CellIsRule(operator="equal", formula=['"REVIEW - NO STANDARD BOX FITS"'], fill=PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")))

autosize(ws, [12, 24, 12, 12, 12, 12, 16, 14, 14, 14, 14, 14, 18, 20, 18, 38])

# ---------- Summary ----------
ws = wb.create_sheet("Summary")
ws.sheet_view.showGridLines = False
ws["B2"] = "Box-Fit Summary"
ws["B2"].font = Font(name=FONT, bold=True, size=16, color="1F4E5F")

labels = [
    ("Total Products Checked", f"=SUMPRODUCT(--('Fit Checker'!A2:A{LAST_ROW}<>\"\"))"),
    ("Products with No Fitting Box", f'=COUNTIF(\'Fit Checker\'!P2:P{LAST_ROW},"REVIEW - NO STANDARD BOX FITS")'),
    ("Total Recommended Packaging Cost (Rs)", f"=SUM('Fit Checker'!N2:N{LAST_ROW})"),
    ("Average Space Utilization (%)", f"=AVERAGEIF('Fit Checker'!P2:P{LAST_ROW},\"OK\",'Fit Checker'!O2:O{LAST_ROW})"),
]
r = 4
for label, formula in labels:
    ws.cell(row=r, column=2, value=label).font = BOLD
    cell = ws.cell(row=r, column=4, value=formula)
    cell.font = BLACK
    if "Rs" in label:
        cell.number_format = "\"Rs\" #,##0"
    if "%" in label:
        cell.number_format = "0.0%"
    r += 1

ws.cell(row=r + 1, column=2, value="Box Usage Breakdown").font = BOLD
zr = r + 3
box_headers = ["Box Code", "Products Assigned"]
for i, h in enumerate(box_headers, start=2):
    ws.cell(row=zr, column=i, value=h)
style_header(ws, zr, len(box_headers), start_col=2)
for i, code in enumerate(["BX1", "BX2", "BX3", "BX4", "BX5", "NO BOX FITS"], start=1):
    rr = zr + i
    ws.cell(row=rr, column=2, value=code).font = BLACK
    ws.cell(row=rr, column=3, value=f"=COUNTIF('Fit Checker'!$M$2:$M${LAST_ROW},B{rr})").font = BLACK
    for c in (2, 3):
        ws.cell(row=rr, column=c).border = BORDER

autosize(ws, [3, 40, 20])
ws.column_dimensions["D"].width = 20

wb.save(r"C:\Users\micro\OneDrive\Documents\excel-automation-templates\volumetric-weight-box-fit-checker\Volumetric_Weight_Box_Fit_Checker.xlsx")
print("saved")
