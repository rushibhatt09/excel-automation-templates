from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

FONT = "Arial"
wb = Workbook()

BLUE = Font(name=FONT, color="0000FF")
BLACK = Font(name=FONT, color="000000")
GREEN = Font(name=FONT, color="008000")
BOLD = Font(name=FONT, bold=True)
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E5F")
YELLOW = PatternFill("solid", fgColor="FFFF00")
TITLE_FILL = PatternFill("solid", fgColor="1F4E5F")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ---------- Sheet: Read Me ----------
ws = wb.active
ws.title = "Read Me"
ws.sheet_view.showGridLines = False
ws["B2"] = "Courier Billing Audit Template"
ws["B2"].font = Font(name=FONT, bold=True, size=16, color="1F4E5F")
ws["B4"] = "What this does"
ws["B4"].font = BOLD
ws["B5"] = ("Compares what your courier actually billed you against what your contracted rate card says you "
            "should have been charged - per shipment - and flags overcharges automatically.")
ws["B7"] = "How to use it"
ws["B7"].font = BOLD
steps = [
    "1. Go to the 'Rate Card' tab and replace the sample zone rates with your own courier contract rates (cells shaded yellow).",
    "2. Go to the 'Shipment Data' tab and paste your own shipment export, replacing the sample rows (keep the same columns).",
    "3. Open the 'Audit' tab - every row recalculates automatically and shows a Verdict: MATCH, OVERCHARGED, or UNDERBILLED.",
    "4. Check the 'Summary Dashboard' tab for total overcharge amount and a zone-wise breakdown - this is your refund claim list.",
]
for i, s in enumerate(steps, start=8):
    ws[f"B{i}"] = s
ws["B13"] = "Notes"
ws["B13"].font = BOLD
ws["B14"] = "- All figures are in INR (Rs). Blue cells = inputs you edit. Black cells = formulas, don't overtype them."
ws["B15"] = "- Chargeable weight = higher of actual weight and volumetric weight (L x W x H / 5000), rounded up to the nearest 0.5 kg slab."
ws["B16"] = "- This is a generic template with illustrative sample data - it is not tied to any specific courier or brand."
for r in range(5, 17):
    ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws[f"B{r}"].font = Font(name=FONT, size=11)
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 100

# ---------- Sheet: Rate Card ----------
ws = wb.create_sheet("Rate Card")
ws.sheet_view.showGridLines = False
ws["A1"] = "Zone Rate Card - SAMPLE VALUES, replace with your own courier contract"
ws["A1"].font = Font(name=FONT, bold=True, size=12, color="1F4E5F")
headers = ["Zone Code", "Zone Name", "Base Weight (kg)", "Base Rate (Rs)",
           "Additional Rate per 0.5kg (Rs)", "COD Charge (% of order value)",
           "COD Minimum Charge (Rs)", "RTO Charge (% of forward freight)"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=3, column=i, value=h)
style_header(ws, 3, len(headers))

rate_rows = [
    ["A", "Local (Same City)", 0.5, 35, 15, 0.02, 30, 1.0],
    ["B", "Regional (Within State)", 0.5, 45, 20, 0.02, 30, 1.0],
    ["C", "Metro to Metro", 0.5, 55, 25, 0.02, 30, 1.0],
    ["D", "Rest of India", 0.5, 65, 30, 0.02, 30, 1.0],
    ["E", "Special (NE / J&K / Ladakh / A&N Islands)", 0.5, 95, 45, 0.02, 30, 1.0],
]
for r, row in enumerate(rate_rows, start=4):
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = BORDER
        if c in (3, 4, 5, 6, 7, 8):
            cell.font = BLUE
            cell.fill = YELLOW
        else:
            cell.font = BLACK
        if c in (6,):
            cell.number_format = "0.0%"
        elif c == 8:
            cell.number_format = "0%"
        elif c in (4, 5, 7):
            cell.number_format = "Rs #,##0"

ws["A11"] = "Blue/yellow cells are sample inputs - overwrite with your own rate card. Add more zone rows if your courier uses a different zone map."
ws["A11"].font = Font(name=FONT, italic=True, size=9, color="666666")
autosize(ws, [10, 38, 16, 14, 24, 22, 20, 24])

# ---------- Sheet: Shipment Data ----------
ws = wb.create_sheet("Shipment Data")
ws.sheet_view.showGridLines = False
sd_headers = ["AWB / Tracking No", "Order ID", "Zone", "Payment Mode", "Order Value (Rs)",
              "Actual Weight (kg)", "Length (cm)", "Width (cm)", "Height (cm)", "Shipment Status",
              "Billed Weight by Courier (kg)", "Billed Amount by Courier (Rs)"]
for i, h in enumerate(sd_headers, start=1):
    ws.cell(row=1, column=i, value=h)
style_header(ws, 1, len(sd_headers))
ws.freeze_panes = "A2"

sample_shipments = [
    ["AWB1001", "ORD5001", "A", "Prepaid", 899, 0.4, 20, 15, 10, "Delivered", 1.0, 75],
    ["AWB1002", "ORD5002", "B", "COD", 1499, 0.8, 25, 20, 12, "Delivered", 1.5, 115],
    ["AWB1003", "ORD5003", "C", "Prepaid", 2199, 1.2, 30, 25, 15, "Delivered", 2.5, 173],
    ["AWB1004", "ORD5004", "D", "COD", 999, 0.3, 15, 15, 10, "RTO", 0.5, 200],
    ["AWB1005", "ORD5005", "E", "Prepaid", 3499, 2.1, 40, 30, 20, "Delivered", 5.0, 560],
    ["AWB1006", "ORD5006", "A", "COD", 599, 0.2, 12, 10, 8, "Delivered", 0.5, 65],
    ["AWB1007", "ORD5007", "B", "Prepaid", 1299, 0.6, 22, 18, 10, "Delivered", 1.0, 80],
    ["AWB1008", "ORD5008", "C", "COD", 1799, 1.0, 28, 22, 14, "RTO", 2.0, 341],
    ["AWB1009", "ORD5009", "D", "Prepaid", 2599, 1.8, 35, 28, 18, "Delivered", 4.0, 305],
    ["AWB1010", "ORD5010", "E", "COD", 4299, 3.0, 45, 35, 25, "Delivered", 8.0, 926],
    ["AWB1011", "ORD5011", "A", "Prepaid", 749, 0.35, 18, 14, 9, "Delivered", 0.5, 55],
    ["AWB1012", "ORD5012", "B", "Prepaid", 1099, 0.5, 20, 16, 11, "Delivered", 1.0, 45],
]
for r, row in enumerate(sample_shipments, start=2):
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = BLUE
        cell.border = BORDER
        if c == 5 or c == 12:
            cell.number_format = "Rs #,##0"
        if c in (6, 7, 8, 9, 11):
            cell.number_format = "0.00"

ws["A16"] = "Replace these sample rows with your own courier billing export. Keep the same column order (paste-special > values if columns differ)."
ws["A16"].font = Font(name=FONT, italic=True, size=9, color="666666")
autosize(ws, [16, 12, 8, 12, 14, 14, 10, 10, 10, 14, 22, 22])

LAST_ROW = 1 + len(sample_shipments)

# ---------- Sheet: Audit ----------
ws = wb.create_sheet("Audit")
ws.sheet_view.showGridLines = False
au_headers = ["AWB / Tracking No", "Zone", "Payment Mode", "Shipment Status", "Actual Weight (kg)",
              "Volumetric Weight (kg)", "Chargeable Weight (kg)", "Billed Weight by Courier (kg)",
              "Weight Check", "Expected Base Freight (Rs)", "Expected Additional Weight Charge (Rs)",
              "Expected COD Charge (Rs)", "Expected RTO Charge (Rs)", "Total Expected Charge (Rs)",
              "Billed Amount by Courier (Rs)", "Variance (Rs)", "Verdict"]
for i, h in enumerate(au_headers, start=1):
    ws.cell(row=1, column=i, value=h)
style_header(ws, 1, len(au_headers))
ws.freeze_panes = "A2"

for i in range(len(sample_shipments)):
    r = i + 2
    sd = i + 2  # row in Shipment Data (same offset)
    ws.cell(row=r, column=1, value=f"='Shipment Data'!A{sd}").font = GREEN
    ws.cell(row=r, column=2, value=f"='Shipment Data'!C{sd}").font = GREEN
    ws.cell(row=r, column=3, value=f"='Shipment Data'!D{sd}").font = GREEN
    ws.cell(row=r, column=4, value=f"='Shipment Data'!J{sd}").font = GREEN
    ws.cell(row=r, column=5, value=f"='Shipment Data'!F{sd}").font = GREEN
    ws.cell(row=r, column=6, value=f"=('Shipment Data'!G{sd}*'Shipment Data'!H{sd}*'Shipment Data'!I{sd})/5000").font = BLACK
    ws.cell(row=r, column=7, value=f"=CEILING(MAX(E{r},F{r}),0.5)").font = BLACK
    ws.cell(row=r, column=8, value=f"='Shipment Data'!K{sd}").font = GREEN
    ws.cell(row=r, column=9, value=f'=IF(H{r}>G{r},"COURIER OVER-ROUNDED WEIGHT","OK")').font = BLACK
    ws.cell(row=r, column=10, value=f"=INDEX('Rate Card'!$D$4:$D$8,MATCH(B{r},'Rate Card'!$A$4:$A$8,0))").font = GREEN
    ws.cell(row=r, column=11, value=(f"=ROUND((G{r}-INDEX('Rate Card'!$C$4:$C$8,MATCH(B{r},'Rate Card'!$A$4:$A$8,0)))/0.5,0)"
                                       f"*INDEX('Rate Card'!$E$4:$E$8,MATCH(B{r},'Rate Card'!$A$4:$A$8,0))")).font = GREEN
    ws.cell(row=r, column=12, value=(f'=IF(C{r}="COD",MAX(\'Shipment Data\'!E{sd}*INDEX(\'Rate Card\'!$F$4:$F$8,MATCH(B{r},\'Rate Card\'!$A$4:$A$8,0)),'
                                       f"INDEX('Rate Card'!$G$4:$G$8,MATCH(B{r},'Rate Card'!$A$4:$A$8,0))),0)")).font = GREEN
    ws.cell(row=r, column=13, value=(f'=IF(D{r}="RTO",(J{r}+K{r})*INDEX(\'Rate Card\'!$H$4:$H$8,MATCH(B{r},\'Rate Card\'!$A$4:$A$8,0)),0)')).font = GREEN
    ws.cell(row=r, column=14, value=f"=J{r}+K{r}+L{r}+M{r}").font = BLACK
    ws.cell(row=r, column=15, value=f"='Shipment Data'!L{sd}").font = GREEN
    ws.cell(row=r, column=16, value=f"=O{r}-N{r}").font = BLACK
    ws.cell(row=r, column=17, value=f'=IF(P{r}>1,"OVERCHARGED",IF(P{r}<-1,"UNDERBILLED","MATCH"))').font = BLACK
    for c in range(1, len(au_headers) + 1):
        ws.cell(row=r, column=c).border = BORDER
    for c in (10, 11, 12, 13, 14, 15, 16):
        ws.cell(row=r, column=c).number_format = "Rs #,##0"
    for c in (5, 6, 7, 8):
        ws.cell(row=r, column=c).number_format = "0.00"

# conditional formatting on Verdict column
verdict_col = "Q"
last = LAST_ROW
ws.conditional_formatting.add(f"{verdict_col}2:{verdict_col}{last}",
    CellIsRule(operator="equal", formula=['"OVERCHARGED"'], fill=PatternFill("solid", fgColor="F8CBAD")))
ws.conditional_formatting.add(f"{verdict_col}2:{verdict_col}{last}",
    CellIsRule(operator="equal", formula=['"UNDERBILLED"'], fill=PatternFill("solid", fgColor="FFE699")))
ws.conditional_formatting.add(f"{verdict_col}2:{verdict_col}{last}",
    CellIsRule(operator="equal", formula=['"MATCH"'], fill=PatternFill("solid", fgColor="C6E0B4")))

autosize(ws, [16, 8, 12, 12, 12, 14, 14, 14, 20, 16, 18, 16, 16, 16, 16, 12, 14])

# ---------- Sheet: Summary Dashboard ----------
ws = wb.create_sheet("Summary Dashboard")
ws.sheet_view.showGridLines = False
ws["B2"] = "Billing Audit Summary"
ws["B2"].font = Font(name=FONT, bold=True, size=16, color="1F4E5F")

labels = [
    ("Total Shipments Audited", f"=COUNTA(Audit!A2:A{LAST_ROW})"),
    ("Total Billed by Courier (Rs)", f"=SUM(Audit!O2:O{LAST_ROW})"),
    ("Total Expected Charge (Rs)", f"=SUM(Audit!N2:N{LAST_ROW})"),
    ("Net Variance (Rs)", f"=SUM(Audit!P2:P{LAST_ROW})"),
    ("Shipments Overcharged", f'=COUNTIF(Audit!Q2:Q{LAST_ROW},"OVERCHARGED")'),
    ("Total Overcharge - Refund Claimable (Rs)", f'=SUMIF(Audit!Q2:Q{LAST_ROW},"OVERCHARGED",Audit!P2:P{LAST_ROW})'),
    ("Shipments Underbilled (courier's loss, FYI)", f'=COUNTIF(Audit!Q2:Q{LAST_ROW},"UNDERBILLED")'),
]
r = 4
for label, formula in labels:
    ws.cell(row=r, column=2, value=label).font = BOLD
    cell = ws.cell(row=r, column=4, value=formula)
    cell.font = BLACK
    if "Rs" in label:
        cell.number_format = "Rs #,##0"
    r += 1

ws.cell(row=r + 1, column=2, value="Zone-wise Breakdown").font = BOLD
zr = r + 3
zone_headers = ["Zone", "Shipments", "Total Billed (Rs)", "Total Expected (Rs)", "Overcharge Amount (Rs)"]
for i, h in enumerate(zone_headers, start=2):
    ws.cell(row=zr, column=i, value=h)
style_header(ws, zr, len(zone_headers), start_col=2)
for i, zone in enumerate(["A", "B", "C", "D", "E"], start=1):
    rr = zr + i
    ws.cell(row=rr, column=2, value=zone).font = BLACK
    ws.cell(row=rr, column=3, value=f'=COUNTIF(Audit!$B$2:$B${LAST_ROW},B{rr})').font = BLACK
    ws.cell(row=rr, column=4, value=f'=SUMIF(Audit!$B$2:$B${LAST_ROW},B{rr},Audit!$O$2:$O${LAST_ROW})').font = BLACK
    ws.cell(row=rr, column=5, value=f'=SUMIF(Audit!$B$2:$B${LAST_ROW},B{rr},Audit!$N$2:$N${LAST_ROW})').font = BLACK
    ws.cell(row=rr, column=6, value=f'=SUMIFS(Audit!$P$2:$P${LAST_ROW},Audit!$B$2:$B${LAST_ROW},B{rr},Audit!$Q$2:$Q${LAST_ROW},"OVERCHARGED")').font = BLACK
    for c in (4, 5, 6):
        ws.cell(row=rr, column=c).number_format = "Rs #,##0"
    for c in range(2, 7):
        ws.cell(row=rr, column=c).border = BORDER

autosize(ws, [3, 34, 12, 20, 14, 14])
ws.column_dimensions["D"].width = 20

wb.save(r"C:\Users\micro\OneDrive\Documents\excel-automation-templates\courier-billing-audit\Courier_Billing_Audit_Template.xlsx")
print("saved")
