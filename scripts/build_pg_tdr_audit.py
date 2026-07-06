from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

FONT = "Arial"
wb = Workbook()

BLUE = Font(name=FONT, color="0000FF")
BLACK = Font(name=FONT, color="000000")
GREEN = Font(name=FONT, color="008000")
BOLD = Font(name=FONT, bold=True)
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E5F", end_color="1F4E5F", fill_type="solid")
YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ---------- Read Me ----------
ws = wb.active
ws.title = "Read Me"
ws.sheet_view.showGridLines = False
ws["B2"] = "Payment Gateway TDR Settlement Audit Template"
ws["B2"].font = Font(name=FONT, bold=True, size=16, color="1F4E5F")
ws["B4"] = "What this does"
ws["B4"].font = BOLD
ws["B5"] = ("Checks every payment gateway settlement against your contracted TDR (Transaction Discount Rate) "
            "card - by payment method - and flags where the gateway deducted more than it should have.")
ws["B7"] = "How to use it"
ws["B7"].font = BOLD
steps = [
    "1. Go to the 'TDR Rate Card' tab and replace the sample rates with your own gateway contract rates (cells shaded yellow).",
    "2. Go to the 'Settlement Data' tab and paste your own settlement export, replacing the sample rows (keep the same columns).",
    "3. Open the 'Audit' tab - every transaction recalculates automatically and shows a Verdict: MATCH, OVERCHARGED, or UNDERBILLED.",
    "4. Check the 'Summary Dashboard' tab for total overcharge amount and a payment-method breakdown - this is your refund claim list.",
]
for i, s in enumerate(steps, start=8):
    ws[f"B{i}"] = s
ws["B13"] = "Notes"
ws["B13"].font = BOLD
ws["B14"] = "- All figures are in INR (Rs). Blue cells = inputs you edit. Black cells = formulas, don't overtype them."
ws["B15"] = "- GST is applied on the TDR amount, not on the transaction amount - this template follows that convention."
ws["B16"] = "- This is a generic template with illustrative sample data - it is not tied to any specific payment gateway or brand."
for r in range(5, 17):
    ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws[f"B{r}"].font = Font(name=FONT, size=11)
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 100

# ---------- TDR Rate Card ----------
ws = wb.create_sheet("TDR Rate Card")
ws.sheet_view.showGridLines = False
ws["A1"] = "TDR Rate Card - SAMPLE VALUES, replace with your own payment gateway contract"
ws["A1"].font = Font(name=FONT, bold=True, size=12, color="1F4E5F")
headers = ["Payment Method", "TDR Rate (% of transaction amount)", "GST on TDR (%)"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=3, column=i, value=h)
style_header(ws, 3, len(headers))

rate_rows = [
    ["UPI", 0.003, 0.18],
    ["Debit Card", 0.004, 0.18],
    ["Credit Card", 0.018, 0.18],
    ["Netbanking", 0.010, 0.18],
    ["Wallet", 0.015, 0.18],
]
for r, row in enumerate(rate_rows, start=4):
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = BORDER
        if c in (2, 3):
            cell.font = BLUE
            cell.fill = YELLOW
            cell.number_format = "0.00%"
        else:
            cell.font = BLACK

ws["A11"] = "Blue/yellow cells are sample inputs - overwrite with your own contracted TDR rates. Add more payment method rows if needed."
ws["A11"].font = Font(name=FONT, italic=True, size=9, color="666666")
autosize(ws, [20, 30, 16])

# ---------- Settlement Data ----------
ws = wb.create_sheet("Settlement Data")
ws.sheet_view.showGridLines = False
sd_headers = ["Transaction ID", "Order ID", "Payment Method", "Transaction Amount (Rs)",
              "TDR Deducted by Gateway (Rs)", "GST on TDR Deducted (Rs)", "Net Settled Amount (Rs)"]
for i, h in enumerate(sd_headers, start=1):
    ws.cell(row=1, column=i, value=h)
style_header(ws, 1, len(sd_headers))
ws.freeze_panes = "A2"

sample_settlements = [
    ["TXN90001", "ORD6001", "UPI", 1299, 10.68, 1.92, 1286.40],
    ["TXN90002", "ORD6002", "Credit Card", 2499, 83.12, 14.96, 2400.92],
    ["TXN90003", "ORD6003", "Debit Card", 899, 3.60, 0.64, 894.76],
    ["TXN90004", "ORD6004", "Netbanking", 1799, 34.94, 6.29, 1757.77],
    ["TXN90005", "ORD6005", "Wallet", 599, 19.15, 3.45, 576.40],
    ["TXN90006", "ORD6006", "UPI", 3499, 10.50, 1.89, 3486.61],
    ["TXN90007", "ORD6007", "Credit Card", 4299, 128.23, 23.08, 4147.69],
    ["TXN90008", "ORD6008", "Debit Card", 1099, 17.11, 3.08, 1078.81],
    ["TXN90009", "ORD6009", "Netbanking", 2199, 43.18, 7.77, 2148.05],
    ["TXN90010", "ORD6010", "Wallet", 1499, 37.74, 6.79, 1454.47],
    ["TXN90011", "ORD6011", "UPI", 999, 2.15, 0.39, 996.46],
    ["TXN90012", "ORD6012", "Credit Card", 5499, 175.25, 31.55, 5292.20],
]
for r, row in enumerate(sample_settlements, start=2):
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = BLUE
        cell.border = BORDER
        if c in (4, 5, 6, 7):
            cell.number_format = "\"Rs\" #,##0.00"

MAX_ROWS = 200  # Audit/Summary formulas are pre-built this far down so extra pasted rows aren't silently dropped
NOTE_ROW = MAX_ROWS + 10  # kept well beyond MAX_ROWS so it never collides with the blank-row guard on column A

ws[f"A{NOTE_ROW}"] = "Replace these sample rows with your own settlement export. Keep the same column order."
ws[f"A{NOTE_ROW}"].font = Font(name=FONT, italic=True, size=9, color="666666")
autosize(ws, [16, 12, 16, 20, 26, 20, 20])

LAST_ROW = 1 + MAX_ROWS

# ---------- Audit ----------
ws = wb.create_sheet("Audit")
ws.sheet_view.showGridLines = False
au_headers = ["Transaction ID", "Payment Method", "Transaction Amount (Rs)", "Expected TDR Rate (%)",
              "Expected TDR Amount (Rs)", "Expected GST on TDR (Rs)", "Expected Total Deduction (Rs)",
              "Actual TDR Deducted (Rs)", "Actual GST Deducted (Rs)", "Actual Total Deduction (Rs)",
              "Variance (Rs)", "Verdict"]
for i, h in enumerate(au_headers, start=1):
    ws.cell(row=1, column=i, value=h)
style_header(ws, 1, len(au_headers))
ws.freeze_panes = "A2"

for i in range(MAX_ROWS):
    r = i + 2
    sd = i + 2
    guard = f"'Settlement Data'!A{sd}=\"\""

    def blank_if_empty(inner):
        return f'=IF({guard},"",{inner})'

    ws.cell(row=r, column=1, value=blank_if_empty(f"'Settlement Data'!A{sd}")).font = GREEN
    ws.cell(row=r, column=2, value=blank_if_empty(f"'Settlement Data'!C{sd}")).font = GREEN
    ws.cell(row=r, column=3, value=blank_if_empty(f"'Settlement Data'!D{sd}")).font = GREEN
    ws.cell(row=r, column=4, value=blank_if_empty(f"INDEX('TDR Rate Card'!$B$4:$B$8,MATCH(B{r},'TDR Rate Card'!$A$4:$A$8,0))")).font = GREEN
    ws.cell(row=r, column=5, value=blank_if_empty(f"C{r}*D{r}")).font = BLACK
    ws.cell(row=r, column=6, value=blank_if_empty(f"E{r}*INDEX('TDR Rate Card'!$C$4:$C$8,MATCH(B{r},'TDR Rate Card'!$A$4:$A$8,0))")).font = GREEN
    ws.cell(row=r, column=7, value=blank_if_empty(f"E{r}+F{r}")).font = BLACK
    ws.cell(row=r, column=8, value=blank_if_empty(f"'Settlement Data'!E{sd}")).font = GREEN
    ws.cell(row=r, column=9, value=blank_if_empty(f"'Settlement Data'!F{sd}")).font = GREEN
    ws.cell(row=r, column=10, value=blank_if_empty(f"H{r}+I{r}")).font = BLACK
    ws.cell(row=r, column=11, value=blank_if_empty(f"J{r}-G{r}")).font = BLACK
    ws.cell(row=r, column=12, value=blank_if_empty(f'IF(K{r}>0.5,"OVERCHARGED",IF(K{r}<-0.5,"UNDERBILLED","MATCH"))')).font = BLACK
    for c in range(1, len(au_headers) + 1):
        ws.cell(row=r, column=c).border = BORDER
    ws.cell(row=r, column=4).number_format = "0.00%"
    for c in (3, 5, 6, 7, 8, 9, 10, 11):
        ws.cell(row=r, column=c).number_format = "\"Rs\" #,##0.00"

verdict_col = "L"
last = LAST_ROW
ws.conditional_formatting.add(f"{verdict_col}2:{verdict_col}{last}",
    CellIsRule(operator="equal", formula=['"OVERCHARGED"'], fill=PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")))
ws.conditional_formatting.add(f"{verdict_col}2:{verdict_col}{last}",
    CellIsRule(operator="equal", formula=['"UNDERBILLED"'], fill=PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")))
ws.conditional_formatting.add(f"{verdict_col}2:{verdict_col}{last}",
    CellIsRule(operator="equal", formula=['"MATCH"'], fill=PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")))

autosize(ws, [16, 16, 18, 16, 18, 18, 20, 18, 18, 18, 14, 16])

# ---------- Summary Dashboard ----------
ws = wb.create_sheet("Summary Dashboard")
ws.sheet_view.showGridLines = False
ws["B2"] = "Settlement Audit Summary"
ws["B2"].font = Font(name=FONT, bold=True, size=16, color="1F4E5F")

labels = [
    ("Total Transactions Audited", f'=SUMPRODUCT(--(Audit!A2:A{LAST_ROW}<>""))'),
    ("Total Transaction Value (Rs)", f"=SUM(Audit!C2:C{LAST_ROW})"),
    ("Total Actual Deduction (Rs)", f"=SUM(Audit!J2:J{LAST_ROW})"),
    ("Total Expected Deduction (Rs)", f"=SUM(Audit!G2:G{LAST_ROW})"),
    ("Net Variance (Rs)", f"=SUM(Audit!K2:K{LAST_ROW})"),
    ("Transactions Overcharged", f'=COUNTIF(Audit!L2:L{LAST_ROW},"OVERCHARGED")'),
    ("Total Overcharge - Refund Claimable (Rs)", f'=SUMIF(Audit!L2:L{LAST_ROW},"OVERCHARGED",Audit!K2:K{LAST_ROW})'),
    ("Transactions Underbilled (gateway's loss, FYI)", f'=COUNTIF(Audit!L2:L{LAST_ROW},"UNDERBILLED")'),
]
r = 4
for label, formula in labels:
    ws.cell(row=r, column=2, value=label).font = BOLD
    cell = ws.cell(row=r, column=4, value=formula)
    cell.font = BLACK
    if "Rs" in label:
        cell.number_format = "\"Rs\" #,##0.00"
    r += 1

ws.cell(row=r + 1, column=2, value="Payment-Method Breakdown").font = BOLD
zr = r + 3
zone_headers = ["Payment Method", "Transactions", "Total Deducted (Rs)", "Total Expected (Rs)", "Overcharge Amount (Rs)"]
for i, h in enumerate(zone_headers, start=2):
    ws.cell(row=zr, column=i, value=h)
style_header(ws, zr, len(zone_headers), start_col=2)
for i, method in enumerate(["UPI", "Debit Card", "Credit Card", "Netbanking", "Wallet"], start=1):
    rr = zr + i
    ws.cell(row=rr, column=2, value=method).font = BLACK
    ws.cell(row=rr, column=3, value=f'=COUNTIF(Audit!$B$2:$B${LAST_ROW},B{rr})').font = BLACK
    ws.cell(row=rr, column=4, value=f'=SUMIF(Audit!$B$2:$B${LAST_ROW},B{rr},Audit!$J$2:$J${LAST_ROW})').font = BLACK
    ws.cell(row=rr, column=5, value=f'=SUMIF(Audit!$B$2:$B${LAST_ROW},B{rr},Audit!$G$2:$G${LAST_ROW})').font = BLACK
    ws.cell(row=rr, column=6, value=f'=SUMIFS(Audit!$K$2:$K${LAST_ROW},Audit!$B$2:$B${LAST_ROW},B{rr},Audit!$L$2:$L${LAST_ROW},"OVERCHARGED")').font = BLACK
    for c in (4, 5, 6):
        ws.cell(row=rr, column=c).number_format = "\"Rs\" #,##0.00"
    for c in range(2, 7):
        ws.cell(row=rr, column=c).border = BORDER

autosize(ws, [3, 40, 14, 20, 20, 20])
ws.column_dimensions["D"].width = 22

wb.save(r"C:\Users\micro\OneDrive\Documents\excel-automation-templates\payment-gateway-tdr-audit\Payment_Gateway_TDR_Audit_Template.xlsx")
print("saved")
